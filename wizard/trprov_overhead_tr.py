# -*- coding: utf-8 -*-

from odoo import api, fields, models
from collections import defaultdict
import base64
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter


class TrprovOverheadTr(models.TransientModel):
    _name = 'trprov.overhead.tr'
    _description = 'Reporte Overhead'

    # Definición de campos para el modelo
    report_from_date = fields.Date(string="Reporte desde", required=True, default=fields.Date.context_today)
    report_to_date = fields.Date(string="Reporte hasta", required=True, default=fields.Date.context_today)
    categ_ids = fields.Many2many('product.category', string="Categoria de los productos")
    res_seller_ids = fields.Many2many('account.analytic.account', string="Cuentas Analíticas", required=True)
    company_id = fields.Many2one(comodel_name="res.company", string="Compañia", required=True,
                                 default=lambda self: self.env.company.id)
    file_content = fields.Binary(string="Archivo Contenido")

    # Acción para generar el archivo Excel y mostrar el enlace de descarga
    def action_generate_excel(self):
        wk_book = openpyxl.Workbook()
        currency_format = NamedStyle(name='currency', number_format='"$"#,##0.00')
        bold_font = Font(bold=True)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

        data = sorted(self.get_data_status_results(), key=lambda x: x['account_type'])

        wk_book.remove(wk_book.active)

        sheets = {}
        last_account_type = None
        separator_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        for item in data:
            analytic_account_name = item['analytic_account_name']
            if analytic_account_name not in sheets:
                sheet = wk_book.create_sheet(title=analytic_account_name)
                sheets[analytic_account_name] = sheet
                last_account_type = None
                report_title = f"Estado de resultados desde {self.report_from_date} hasta {self.report_to_date}"
                sheet.append(["", "", report_title])
                for cell in sheet["B2:F2"]:
                    cell[0].font = bold_font

                company_name = self.env.company.name
                sheet.append(["", "Compañía:", company_name])
                for cell in sheet["B3:F3"]:
                    cell[0].font = bold_font

                sheet.append(["", "Nombre Cto Costo:", analytic_account_name])
                for cell in sheet["B4:F4"]:
                    cell[0].font = bold_font

                # Deja una fila vacía
                sheet.append([])
                headers = [
                    "Concepto", "Nombre Cto. Costo", "Nombre Cta Agrupador", "Enero", "Febrero",
                    "Marzo",
                    "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
                    "Total Resultado"
                ]
                sheet.append(headers)
                for col_index, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=6, column=col_index)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
            else:
                sheet = sheets[analytic_account_name]

            if last_account_type and item['account_type'] != last_account_type:
                row_index = sheet.max_row + 1
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_index, column=col)
                    cell.fill = separator_fill

            last_account_type = item['account_type']

            row_index = sheet.max_row + 1

            sheet.cell(row=row_index, column=1, value=item['account_type'])
            sheet.cell(row=row_index, column=2, value=item['analytic_account_name'])
            sheet.cell(row=row_index, column=3, value=item['financial_account_name'])

            for month_index in range(1, 13):
                month_value = item.get(f'month_{month_index}', 0)
                sheet.cell(row=row_index, column=month_index + 3, value=month_value).style = currency_format
            sheet.cell(row=row_index, column=16, value=item['total_result']).style = currency_format

        # Después de imprimir los datos mensuales, agregar la fila de sumas totales
        for month_index in range(1, 13):
            month_total = sum(item.get(f'month_{month_index}', 0) for item in data)
            sheet.cell(row=row_index + 1, column=month_index + 3, value=month_total).style = currency_format

        # Calcular la suma total para la columna "Total Resultado"
        total_result_sum = sum(item['total_result'] for item in data)
        sheet.cell(row=row_index + 1, column=16, value=total_result_sum).style = currency_format

        output = BytesIO()
        wk_book.save(output)
        output.seek(0)
        base64_content = base64.b64encode(output.read())
        output.close()

        self.write({'file_content': base64_content})

        # Crear formato para nombre de salida de archivo
        formatted_from_date = self.report_from_date.strftime('%Y-%m-%d')
        formatted_to_date = self.report_to_date.strftime('%Y-%m-%d')
        filename = f"Estado_resultados_{formatted_from_date}_{formatted_to_date}.xlsx"

        return {
            'type': 'ir.actions.act_url',
            'url': f"web/content/?model={self._name}&id={self.id}&field=file_content&filename={filename}&download=true",
            'target': 'self',
        }

    # Función para obtener datos del estado de resultados desde Apuntes contables
    def get_data_status_results(self):
        ACCOUNT_TYPE_MAPPING = {
            "asset_receivable": "Por cobrar",
            "asset_cash": "Banco y efectivo",
            "asset_current": "Activos Circulantes",
            "asset_non_current": "Activos no-circulantes",
            "asset_prepayments": "Prepagos",
            "asset_fixed": "Activos Fijos",
            "liability_payable": "Por pagar",
            "liability_credit_card": "Tarjeta de Crédito",
            "liability_current": "Pasivos Circulantes",
            "liability_non_current": "Pasivos no-circulantes",
            "equity": "Capital",
            "equity_unaffected": "Ganancias del año actual",
            "income": "Ingreso",
            "income_other": "Otro Ingreso",
            "expense": "Gastos",
            "expense_depreciation": "Depreciación",
            "expense_direct_cost": "Costo de ingresos",
            "off_balance": "Hoja fuera de balance",
        }

        final_data = []
        for rec in self:
            data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
            account_info = defaultdict(dict)

            analytic_lines = rec.env['account.analytic.line'].search([
                ('account_id', 'in', rec.res_seller_ids.ids),
                ('date', '>=', rec.report_from_date),
                ('date', '<=', rec.report_to_date),
            ])

            for line in analytic_lines:
                analytic_account_name = line.account_id.name
                financial_account_name = line.general_account_id.name
                financial_account_type = line.trprovwi_general_account_type_tr
                analytic_account_id = line.account_id.id
                financial_account_id = line.general_account_id.id

                account_info[analytic_account_id]['name'] = analytic_account_name
                account_info[analytic_account_id][financial_account_id] = {
                    'name': financial_account_name,
                    'type': financial_account_type
                }

                for month in range(1, 13):
                    if line.date.month == month:
                        amount = line.amount or 0
                        data[analytic_account_id][financial_account_id][f'month_{month}'] += amount
                        data[analytic_account_id][financial_account_id]['total_result'] += amount

            for analytic_id, financial_accounts in data.items():
                for financial_id, months in financial_accounts.items():
                    account_type_key = account_info[analytic_id][financial_id]['type']
                    account_type_value = ACCOUNT_TYPE_MAPPING.get(account_type_key,
                                                                  account_type_key)
                    final_data.append({
                        'analytic_account_id': analytic_id,
                        'analytic_account_name': account_info[analytic_id]['name'],
                        'financial_account_id': financial_id,
                        'financial_account_name': account_info[analytic_id][financial_id]['name'],
                        'account_type': account_type_value,
                        **months
                    })

        return final_data